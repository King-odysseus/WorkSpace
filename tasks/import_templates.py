"""Downloadable import templates for the frontend and reporting agents."""

import csv
import io

from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET

from .views import require_workspace_member


TEMPLATE_COLUMNS = {
    'tasks': ['Code', 'Title', 'Description', 'Owner', 'Supporters', 'Project', 'Workstream', 'Phase', 'Bucket', 'Priority', 'Status', 'Start date', 'Due date', 'Progress %', 'Labels', 'Blocker details', 'Actual completion'],
    'projects': ['Name', 'Description', 'Status', 'Start date', 'End date', 'Due date', 'Timezone', 'Week anchor date', 'Due soon days', 'Configuration JSON'],
    'stakeholders': ['Project', 'Name', 'Role', 'Email', 'Influence', 'Interest', 'Notes'],
}


@require_GET
def import_template(request, workspace_id, kind, file_format):
    _, error = require_workspace_member(request, workspace_id)
    if error:
        return error
    kind = kind.lower()
    file_format = file_format.lower()
    if kind not in TEMPLATE_COLUMNS:
        return JsonResponse({'error': 'Template kind must be tasks, projects, or stakeholders.'}, status=404)
    if file_format == 'csv':
        output = io.StringIO(newline='')
        writer = csv.writer(output)
        writer.writerow(TEMPLATE_COLUMNS[kind])
        writer.writerow(_example_row(kind))
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    elif file_format == 'xlsx':
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'General' if kind == 'tasks' else kind.title()
        sheet.append(TEMPLATE_COLUMNS[kind])
        sheet.append(_example_row(kind))
        if kind == 'tasks':
            instructions = workbook.create_sheet('Instructions')
            instructions.append(['Column', 'Guidance'])
            instructions.append(['Code', 'Optional for new tasks; existing codes are never reused.'])
            instructions.append(['Owner', 'Workspace member email or unique full name.'])
            instructions.append(['Status', 'todo, in_progress, blocked, review, on_hold, cancelled, or done.'])
            instructions.append(['Project', 'Existing project name; create projects first when importing separately.'])
        elif kind == 'projects':
            instructions = workbook.create_sheet('Instructions')
            instructions.append(['Status', 'planning, active, paused, or completed'])
            instructions.append(['Configuration JSON', 'Optional JSON object for project-specific settings.'])
        else:
            instructions = workbook.create_sheet('Instructions')
            instructions.append(['Influence / Interest', 'low, medium, or high'])
        buffer = io.BytesIO()
        workbook.save(buffer)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        return JsonResponse({'error': 'Template format must be csv or xlsx.'}, status=400)
    response['Content-Disposition'] = f'attachment; filename="workspace-{kind}-import-template.{file_format}"'
    return response


def _example_row(kind):
    if kind == 'tasks':
        return ['TASK-EXAMPLE', 'Example task', 'Replace this example row', '', '', '', '', '', 'Backlog', 'normal', 'todo', '', '2026-12-31', '0', '', '', '']
    if kind == 'projects':
        return ['Example project', 'Replace this example row', 'planning', '2026-01-01', '2026-03-31', '2026-03-31', 'Europe/London', '2026-01-05', '7', '{}']
    return ['Example project', 'Example stakeholder', 'Sponsor', 'stakeholder@example.com', 'medium', 'medium', 'Replace this example row']
