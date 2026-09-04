"""Excel migration framework.

Imports tasks from a workbook with three sheet roles:

* **General**  - the authoritative task source (one row per task). Upsert keyed by
  the task ``code`` when present; otherwise a new task is created.
* **Lists**    - configuration (allowed buckets, priorities, workstreams, phases,
  statuses). Read as a header row of kinds with a column of values each.
* **Owner sheets** - every other sheet, keyed by task ``code``, used ONLY for
  execution enrichment (progress, status, blocker details, actual completion) of
  tasks that already exist. Owner sheets never create tasks.

Design guarantees:

* **Preview before commit** - ``build_import_plan`` performs every validation and
  user match without writing anything; ``commit_import_plan`` then applies the
  plan inside a single ``transaction.atomic()`` block.
* **User matching** - owners/supporters are matched to existing users by email
  (case-insensitive) then by unique full name; unmatched emails become pending
  invitations, unmatched names become exceptions.
* **Exception reporting** - every row-level problem is collected with its row
  number and field so the caller can show a precise report.
* **Transactional** - commit is all-or-nothing; any unexpected error rolls the
  whole import back.

WORKBOOK-REQUIRED NOTES (cannot be verified without the original file):
  * the exact General-sheet column headers (override via ``column_map``);
  * the exact Lists-sheet layout and the owner-sheet enrichment columns.
"""

import logging
from datetime import date, datetime

from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify

from .models import ImportRun, Membership, Project, Task, TaskChangeHistory, TaskCodeRegistry, Workspace, WorkspaceInvitation

STATUS_ALIASES = {
    'todo': 'todo', 'to do': 'todo', 'not started': 'todo', 'backlog': 'todo',
    'in progress': 'in_progress', 'in_progress': 'in_progress', 'doing': 'in_progress', 'active': 'in_progress',
    'blocked': 'blocked', 'review': 'review', 'in review': 'review',
    'on hold': 'on_hold', 'on_hold': 'on_hold', 'paused': 'on_hold',
    'cancelled': 'cancelled', 'canceled': 'cancelled',
    'done': 'done', 'complete': 'done', 'completed': 'done',
}
PRIORITY_ALIASES = {
    'urgent': 'urgent', 'high': 'high', 'normal': 'normal', 'medium': 'normal', 'low': 'low',
}

# Default column headers for the General sheet. Overridable per import; the
# defaults are a starting point that MUST be verified against the real workbook.
DEFAULT_GENERAL_COLUMNS = {
    'code': 'Code',
    'title': 'Title',
    'description': 'Description',
    'owner': 'Owner',
    'supporters': 'Supporters',
    'project': 'Project',
    'workstream': 'Workstream',
    'phase': 'Phase',
    'bucket': 'Bucket',
    'priority': 'Priority',
    'status': 'Status',
    'start_date': 'Start date',
    'due_date': 'Due date',
    'progress_percent': 'Progress %',
    'labels': 'Labels',
    'blocker_details': 'Blocker details',
    'actual_completion_date': 'Actual completion',
}

GENERAL_SHEET = 'General'
LISTS_SHEET = 'Lists'

# Canonical keys that map onto Task fields (used for enrichment sheets too).
TASK_FIELDS = {
    'title', 'code', 'description', 'project', 'workstream', 'phase', 'bucket',
    'priority', 'status', 'start_date', 'due_date', 'progress_percent', 'labels',
    'blocker_details', 'actual_completion_date',
}


def _normalize_status(value):
    return STATUS_ALIASES.get(str(value or '').strip().lower())


def _normalize_priority(value):
    return PRIORITY_ALIASES.get(str(value or '').strip().lower())


def _parse_date(value):
    if value in (None, ''):
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    text = str(value).strip()
    try:
        return date.fromisoformat(text), None
    except ValueError:
        return None, f'Invalid date "{text}" (expected YYYY-MM-DD).'


def _split_list(value):
    if value in (None, ''):
        return []
    if isinstance(value, (list, tuple)):
        return [str(item).strip() for item in value if str(item).strip()]
    return [part.strip() for part in str(value).replace(';', ',').split(',') if part.strip()]


def _column_map_from(header_row, defaults):
    """Map canonical keys to column indexes from the General sheet header row."""
    header = [str(cell or '').strip() for cell in header_row]
    mapping = {}
    for key, expected in defaults.items():
        if expected in header:
            mapping[key] = header.index(expected)
        else:
            for idx, value in enumerate(header):
                if value.lower() == expected.lower():
                    mapping[key] = idx
                    break
    return mapping, header


def _member_indexes(workspace):
    members = list(Membership.objects.filter(workspace=workspace).select_related('user'))
    email_index = {}
    name_index = {}
    for membership in members:
        user = membership.user
        email_index[user.email.lower()] = user
        full_name = f'{user.first_name} {user.last_name}'.strip().lower()
        if full_name:
            name_index.setdefault(full_name, []).append(user)
    return email_index, name_index, members


def _match_user(value, email_index, name_index):
    """Return (user, invite_email) for an owner/supporter value.

    user is None when unmatched; invite_email is set only when the value is an
    email that can be invited (otherwise an unmatched name is an exception).
    """
    value = str(value or '').strip()
    if not value:
        return None, None
    lowered = value.lower()
    if '@' in lowered:
        return email_index.get(lowered), lowered
    matches = name_index.get(lowered, [])
    if len(matches) == 1:
        return matches[0], None
    return None, None


def _build_invitation_role():
    return 'member'


def build_import_plan(workspace, workbook, column_map=None):
    """Parse and validate a workbook into an import plan (no writes)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError('openpyxl is required for Excel imports (pip install openpyxl).')

    column_map = column_map or DEFAULT_GENERAL_COLUMNS
    wb = load_workbook(workbook, data_only=True, read_only=True)
    exceptions = []
    invitations = []
    normalized = []
    lists = {}
    owner_sheets = {}

    try:
        general_ws = wb[GENERAL_SHEET]
    except KeyError:
        return {'error': f'A "{GENERAL_SHEET}" sheet is required as the authoritative task source.'}

    rows = list(general_ws.iter_rows(values_only=True))
    if not rows:
        return {'error': 'The General sheet is empty.'}

    mapping, _header = _column_map_from(rows[0], column_map)
    if 'title' not in mapping:
        return {'error': 'The General sheet must include a Title column.'}

    email_index, name_index, _members = _member_indexes(workspace)
    existing_by_code = {}
    for task in Task.objects.filter(workspace=workspace).exclude(code=''):
        existing_by_code.setdefault(task.code.strip().lower(), []).append(task)
    reserved_codes = set(TaskCodeRegistry.objects.filter(workspace=workspace).values_list('code', flat=True))
    reserved_codes = {code.strip().lower() for code in reserved_codes}

    seen_codes = set()
    for row_number, row in enumerate(rows[1:], start=2):
        if row is None or all(cell in (None, '') for cell in row):
            continue
        values = {key: (row[idx] if idx < len(row) else None) for key, idx in mapping.items()}

        title = str(values.get('title') or '').strip()
        if not title:
            exceptions.append({'row': row_number, 'field': 'title', 'message': 'Title is required.'})
            continue

        code = str(values.get('code') or '').strip()
        if code:
            code_lower = code.lower()
            if code_lower in seen_codes:
                exceptions.append({'row': row_number, 'field': 'code', 'message': f'Duplicate code "{code}" in file.'})
                continue
            seen_codes.add(code_lower)

        status = _normalize_status(values.get('status'))
        if values.get('status') not in (None, '') and status is None:
            exceptions.append({'row': row_number, 'field': 'status', 'message': f'Unknown status "{values.get("status")}".'})
            continue

        priority = _normalize_priority(values.get('priority'))
        if values.get('priority') not in (None, '') and priority is None:
            exceptions.append({'row': row_number, 'field': 'priority', 'message': f'Unknown priority "{values.get("priority")}".'})
            continue

        start_date, start_error = _parse_date(values.get('start_date'))
        if start_error:
            exceptions.append({'row': row_number, 'field': 'start_date', 'message': start_error})
            continue
        due_date, due_error = _parse_date(values.get('due_date'))
        if due_error:
            exceptions.append({'row': row_number, 'field': 'due_date', 'message': due_error})
            continue
        actual_completion_date, actual_error = _parse_date(values.get('actual_completion_date'))
        if actual_error:
            exceptions.append({'row': row_number, 'field': 'actual_completion_date', 'message': actual_error})
            continue
        if start_date and due_date and due_date < start_date:
            exceptions.append({'row': row_number, 'field': 'due_date', 'message': 'Target date cannot precede start date.'})
            continue
        if start_date and actual_completion_date and actual_completion_date < start_date:
            exceptions.append({'row': row_number, 'field': 'actual_completion_date', 'message': 'Actual completion cannot precede start date.'})
            continue
        if actual_completion_date and status != 'done':
            exceptions.append({'row': row_number, 'field': 'actual_completion_date', 'message': 'Actual completion requires completed status.'})
            continue

        progress_raw = values.get('progress_percent')
        progress = None
        if progress_raw not in (None, ''):
            try:
                progress = int(float(progress_raw))
                if not 0 <= progress <= 100:
                    raise ValueError
            except (TypeError, ValueError):
                exceptions.append({'row': row_number, 'field': 'progress_percent', 'message': f'Invalid progress "{progress_raw}".'})
                continue
        blocker_details = str(values.get('blocker_details') or '').strip()
        if status == 'blocked' and not blocker_details:
            exceptions.append({'row': row_number, 'field': 'blocker_details', 'message': 'Blocked tasks require blocker details.'})
            continue

        owner_value = values.get('owner')
        owner, owner_invite = _match_user(owner_value, email_index, name_index)
        if owner_value and owner is None and owner_invite is None:
            exceptions.append({'row': row_number, 'field': 'owner', 'message': f'Could not match owner "{owner_value}".'})
            continue
        if owner_invite:
            invitations.append({'email': owner_invite, 'role': _build_invitation_role()})

        supporter_users = []
        supporter_invites = []
        for supporter_value in _split_list(values.get('supporters')):
            supporter, invite = _match_user(supporter_value, email_index, name_index)
            if supporter:
                supporter_users.append(supporter)
            elif invite:
                supporter_invites.append(invite)
            else:
                exceptions.append({'row': row_number, 'field': 'supporters', 'message': f'Could not match supporter "{supporter_value}".'})
        for invite in supporter_invites:
            invitations.append({'email': invite, 'role': _build_invitation_role()})

        project_ref = None
        project_name = str(values.get('project') or '').strip()
        if project_name:
            project_ref = Project.objects.filter(workspace=workspace, name=project_name).first()
            if project_ref is None:
                exceptions.append({'row': row_number, 'field': 'project', 'message': f'Unknown project "{project_name}" (create it first, or add it to the Lists sheet).'})
                continue

        matches = existing_by_code.get(code_lower, []) if code else []
        if len(matches) > 1:
            exceptions.append({'row': row_number, 'field': 'code', 'message': f'Code "{code}" matches multiple existing tasks.'})
            continue
        existing = matches[0] if matches else None
        if code and existing is None and code_lower in reserved_codes:
            exceptions.append({'row': row_number, 'field': 'code', 'message': f'Code "{code}" was previously used and cannot be reused.'})
            continue
        normalized.append({
            'row': row_number,
            'action': 'update' if existing else 'create',
            'existing_task_id': existing.id if existing else None,
            'code': code,
            'title': title,
            'description': str(values.get('description') or '').strip(),
            'assignee': owner,
            'supporters': supporter_users,
            'project_ref': project_ref,
            'workstream': str(values.get('workstream') or '').strip(),
            'phase': str(values.get('phase') or '').strip(),
            'bucket': str(values.get('bucket') or 'Backlog').strip() or 'Backlog',
            'priority': priority or 'normal',
            'status': status or 'todo',
            'start_date': start_date,
            'due_date': due_date,
            'actual_completion_date': actual_completion_date,
            'progress_percent': progress,
            'blocker_details': blocker_details,
            'labels': _split_list(values.get('labels')),
        })

    # Lists sheet (configuration).
    if LISTS_SHEET in wb.sheetnames:
        lists = _parse_lists_sheet(wb[LISTS_SHEET])

    # Owner sheets (execution enrichment only).
    for sheet_name in wb.sheetnames:
        if sheet_name in (GENERAL_SHEET, LISTS_SHEET):
            continue
        owner_sheets[sheet_name] = _parse_owner_sheet(wb[sheet_name], {key: values[0] for key, values in existing_by_code.items() if len(values) == 1})

    summary = {
        'total_rows': len(normalized) + len(exceptions),
        'creates': sum(1 for n in normalized if n['action'] == 'create'),
        'updates': sum(1 for n in normalized if n['action'] == 'update'),
        'exceptions': len(exceptions),
        'invitations': len(invitations),
    }

    return {
        'source': getattr(workbook, 'name', ''),
        'column_map': {key: value for key, value in column_map.items()},
        'lists': lists,
        'owner_sheets': {name: rows for name, rows in owner_sheets.items()},
        'normalized': normalized,
        'exceptions': exceptions,
        'invitations': invitations,
        'summary': summary,
    }


def _parse_lists_sheet(worksheet):
    """Read a Lists sheet as a header row of kinds + a column of values each."""
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(cell or '').strip() for cell in rows[0]]
    config = {}
    for row in rows[1:]:
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            if value in (None, ''):
                continue
            config.setdefault(header.lower(), []).append(str(value).strip())
    return config


def _parse_owner_sheet(worksheet, existing_by_code):
    """Read an owner sheet as code-keyed enrichment rows (no task creation)."""
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or '').strip() for cell in rows[0]]
    enrichment = []
    for row in rows[1:]:
        values = dict(zip(headers, row))
        code = str(values.get('Code') or values.get('code') or '').strip()
        if not code:
            continue
        task = existing_by_code.get(code.lower())
        enrichment.append({
            'code': code,
            'task_id': task.id if task else None,
            'progress_percent': values.get('Progress %') or values.get('progress_percent'),
            'status': values.get('Status') or values.get('status'),
            'blocker_details': values.get('Blocker details') or values.get('blocker_details'),
            'actual_completion_date': values.get('Actual completion') or values.get('actual_completion_date'),
        })
    return enrichment


def commit_import_plan(workspace, actor, plan):
    """Apply a validated import plan transactionally."""
    if plan.get('error'):
        raise ValueError(plan['error'])

    created = 0
    updated = 0
    with transaction.atomic():
        touched = {}
        previous = {}
        for norm in plan['normalized']:
            task, is_new = _upsert_task(workspace, norm)
            touched[task.id] = task
            previous[task.id] = None if is_new else getattr(task, '_import_previous', None)
            if is_new:
                created += 1
            else:
                updated += 1
            task.supporters.set(norm['supporters'])

        for enrichment_sheet in plan['owner_sheets'].values():
            for entry in enrichment_sheet:
                if not entry.get('task_id'):
                    continue
                task = Task.objects.filter(id=entry['task_id'], workspace_id=workspace.id).first()
                if task is None:
                    continue
                if task.id not in previous:
                    previous[task.id] = _history_snapshot(task)
                    _ensure_code_registry(workspace, task.code, task.id)
                _apply_enrichment(task, entry)
                touched[task.id] = task

        for invite in plan['invitations']:
            WorkspaceInvitation.objects.get_or_create(
                workspace=workspace,
                email=invite['email'],
                status='pending',
                defaults={'role': invite['role'], 'invited_by': actor},
            )

        for task in touched.values():
            if task.status == 'done':
                task.progress_percent = 100
                if task.actual_completion_date:
                    task.completed_at = timezone.make_aware(datetime.combine(task.actual_completion_date, datetime.min.time()))
            elif task.actual_completion_date:
                raise ValueError(f'Task "{task.code}" has an actual completion date but is not completed.')
            task.full_clean()
            task.save()
            TaskChangeHistory.objects.create(
                task=task, task_code=task.code, workspace=workspace, actor=actor,
                field='imported' if previous.get(task.id) is None else 'import_updated',
                previous_value=previous.get(task.id), new_value=_history_snapshot(task),
            )

        ImportRun.objects.create(
            workspace=workspace,
            actor=actor,
            mode='commit',
            source=plan.get('source', ''),
            summary=plan['summary'],
            exceptions=plan['exceptions'],
        )

    return {'created': created, 'updated': updated, 'invitations': len(plan['invitations']), 'exceptions': plan['exceptions']}


def _upsert_task(workspace, norm):
    if norm['action'] == 'update':
        task = Task.objects.get(id=norm['existing_task_id'])
        task._import_previous = _history_snapshot(task)
        _ensure_code_registry(workspace, task.code, task.id)
        _apply_task_fields(task, norm)
        return task, False

    code = norm['code'].strip() if norm['code'] else _reserve_generated_code(workspace)
    if code and (TaskCodeRegistry.objects.filter(workspace=workspace, code__iexact=code).exists() or Task.objects.filter(workspace=workspace, code__iexact=code).exists()):
        raise ValueError(f'Task code "{code}" was previously used and cannot be reused.')
    task = Task.objects.create(workspace=workspace, title=norm['title'], code=code)
    TaskCodeRegistry.objects.create(workspace=workspace, code=code, task_id=task.id)
    norm = {**norm, 'code': code}
    _apply_task_fields(task, norm)
    return task, True


def _reserve_generated_code(workspace):
    locked = Workspace.objects.select_for_update().get(id=workspace.id)
    prefix = (slugify(locked.slug or locked.name).replace('-', '')[:8] or 'TASK').upper()
    number = locked.next_task_number
    while True:
        code = f'{prefix}-{number:06d}'
        if not TaskCodeRegistry.objects.filter(workspace=locked, code__iexact=code).exists() and not Task.objects.filter(workspace=locked, code__iexact=code).exists():
            break
        number += 1
    locked.next_task_number = number + 1
    locked.save(update_fields=['next_task_number'])
    return code


def _ensure_code_registry(workspace, code, task_id):
    registry = TaskCodeRegistry.objects.filter(workspace=workspace, code__iexact=code).first()
    if registry and registry.task_id not in (None, task_id):
        raise ValueError(f'Task code "{code}" is reserved for another task.')
    if registry is None:
        TaskCodeRegistry.objects.create(workspace=workspace, code=code, task_id=task_id)
    elif registry.task_id is None:
        registry.task_id = task_id
        registry.save(update_fields=['task_id'])


def _history_snapshot(task):
    return {
        'title': task.title, 'code': task.code, 'assignee_id': task.assignee_id,
        'project_id': task.project_ref_id, 'status': task.status, 'priority': task.priority,
        'start_date': task.start_date.isoformat() if task.start_date else None,
        'due_date': task.due_date.isoformat() if task.due_date else None,
        'actual_completion_date': task.actual_completion_date.isoformat() if task.actual_completion_date else None,
        'progress_percent': task.progress_percent, 'blocker_details': task.blocker_details,
        'workstream': task.workstream, 'phase': task.phase, 'state': task.state,
        'supporter_ids': list(task.supporters.values_list('id', flat=True)),
    }


def _apply_task_fields(task, norm):
    task.code = norm['code']
    task.title = norm['title']
    task.description = norm['description']
    task.assignee = norm['assignee']
    task.project_ref = norm['project_ref']
    task.workstream = norm['workstream']
    task.phase = norm['phase']
    task.bucket = norm['bucket']
    task.priority = norm['priority']
    task.status = norm['status']
    task.start_date = norm['start_date']
    task.due_date = norm['due_date']
    task.actual_completion_date = norm.get('actual_completion_date')
    task.blocker_details = norm.get('blocker_details', '')
    if norm['progress_percent'] is not None:
        task.progress_percent = norm['progress_percent']
    task.labels = norm['labels']
    if norm['status'] == 'done':
        task.progress_percent = 100
        if task.actual_completion_date:
            task.completed_at = timezone.make_aware(datetime.combine(task.actual_completion_date, datetime.min.time()))
    else:
        task.completed_at = None
    task.save()


def _apply_enrichment(task, entry):
    changed = False
    status = _normalize_status(entry.get('status'))
    if status:
        task.status = status
        changed = True
    progress = entry.get('progress_percent')
    if progress not in (None, ''):
        try:
            task.progress_percent = int(float(progress))
            changed = True
        except (TypeError, ValueError):
            logging.getLogger(__name__).warning(
                'Import: could not parse progress_percent %r for task %s', progress, task.code
            )
    blocker_details = entry.get('blocker_details')
    if blocker_details not in (None, ''):
        task.blocker_details = str(blocker_details).strip()
        changed = True
    actual = _parse_date(entry.get('actual_completion_date'))[0]
    if actual:
        task.actual_completion_date = actual
        changed = True
    if changed:
        task.save()
