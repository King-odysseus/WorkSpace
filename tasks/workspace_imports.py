"""CSV/XLSX project and stakeholder import plans."""

import csv
import io
import json
from datetime import date

from django.db import transaction

from .models import Project, ProjectStakeholder


PROJECT_COLUMNS = {
    'name': 'Name', 'description': 'Description', 'status': 'Status', 'start_date': 'Start date',
    'end_date': 'End date', 'due_date': 'Due date', 'timezone': 'Timezone',
    'week_anchor_date': 'Week anchor date', 'due_soon_days': 'Due soon days', 'configuration': 'Configuration JSON',
}
STAKEHOLDER_COLUMNS = {
    'project': 'Project', 'name': 'Name', 'role': 'Role', 'email': 'Email',
    'influence': 'Influence', 'interest': 'Interest', 'notes': 'Notes',
}


def _rows_from_file(content, filename, sheet_name, columns):
    if filename.lower().endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        return [{key: row.get(header, '') for key, header in columns.items()} for row in reader]
    from openpyxl import load_workbook
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(value or '').strip() for value in values[0]]
    indexes = {}
    for key, expected in columns.items():
        indexes[key] = next((index for index, header in enumerate(headers) if header.lower() == expected.lower()), None)
    return [{key: (row[index] if index is not None and index < len(row) else '') for key, index in indexes.items()} for row in values[1:] if any(value not in (None, '') for value in row)]


def _parse_date(value, field, row, exceptions):
    if value in (None, ''):
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        exceptions.append({'row': row, 'field': field, 'message': f'{field} must use YYYY-MM-DD format.'})
        return None


def build_project_plan(workspace, content, filename):
    rows = _rows_from_file(content, filename, 'Projects', PROJECT_COLUMNS)
    exceptions, normalized = [], []
    existing = {}
    for project in Project.objects.filter(workspace=workspace):
        existing.setdefault(project.name.strip().lower(), []).append(project)
    seen = set()
    for row_number, values in enumerate(rows, start=2):
        name = str(values.get('name') or '').strip()
        if not name or len(name) > 160:
            exceptions.append({'row': row_number, 'field': 'name', 'message': 'Project name is required and must be 160 characters or fewer.'})
            continue
        key = name.lower()
        if key in seen:
            exceptions.append({'row': row_number, 'field': 'name', 'message': f'Duplicate project name "{name}" in file.'})
            continue
        seen.add(key)
        matches = existing.get(key, [])
        if len(matches) > 1:
            exceptions.append({'row': row_number, 'field': 'name', 'message': f'Project name "{name}" matches multiple existing projects.'})
            continue
        parsed = {field: _parse_date(values.get(field), PROJECT_COLUMNS[field], row_number, exceptions) for field in ('start_date', 'end_date', 'due_date', 'week_anchor_date')}
        if parsed['start_date'] and parsed['end_date'] and parsed['end_date'] < parsed['start_date']:
            exceptions.append({'row': row_number, 'field': 'end_date', 'message': 'End date cannot precede start date.'})
            continue
        if parsed['start_date'] and parsed['due_date'] and parsed['due_date'] < parsed['start_date']:
            exceptions.append({'row': row_number, 'field': 'due_date', 'message': 'Due date cannot precede start date.'})
            continue
        status = str(values.get('status') or 'planning').strip().lower()
        if status not in dict(Project.STATUS_CHOICES):
            exceptions.append({'row': row_number, 'field': 'status', 'message': f'Unknown project status "{status}".'})
            continue
        try:
            due_soon_days = int(values.get('due_soon_days') or 7)
            if not 0 <= due_soon_days <= 365:
                raise ValueError
        except (TypeError, ValueError):
            exceptions.append({'row': row_number, 'field': 'due_soon_days', 'message': 'Due soon days must be between 0 and 365.'})
            continue
        configuration = {}
        raw_configuration = values.get('configuration')
        if raw_configuration not in (None, ''):
            try:
                configuration = json.loads(str(raw_configuration))
                if not isinstance(configuration, dict):
                    raise ValueError
            except (TypeError, ValueError, json.JSONDecodeError):
                exceptions.append({'row': row_number, 'field': 'configuration', 'message': 'Configuration JSON must be a JSON object.'})
                continue
        normalized.append({'row': row_number, 'action': 'update' if matches else 'create', 'existing_id': matches[0].id if matches else None, 'name': name, 'description': str(values.get('description') or '').strip(), 'status': status, **parsed, 'timezone': str(values.get('timezone') or '').strip(), 'due_soon_days': due_soon_days, 'configuration': configuration})
    return {'kind': 'projects', 'summary': {'total_rows': len(rows), 'creates': sum(item['action'] == 'create' for item in normalized), 'updates': sum(item['action'] == 'update' for item in normalized), 'exceptions': len(exceptions)}, 'normalized': normalized, 'exceptions': exceptions}


def commit_project_plan(workspace, actor, plan):
    with transaction.atomic():
        for item in plan['normalized']:
            project = Project.objects.filter(id=item['existing_id'], workspace=workspace).first() if item['existing_id'] else Project(workspace=workspace)
            if project is None:
                raise ValueError(f'Project {item["name"]} no longer exists.')
            for field in ('name', 'description', 'status', 'start_date', 'end_date', 'due_date', 'timezone', 'week_anchor_date', 'due_soon_days', 'configuration'):
                setattr(project, field, item[field])
            project.full_clean()
            project.save()
    return {'created': sum(item['action'] == 'create' for item in plan['normalized']), 'updated': sum(item['action'] == 'update' for item in plan['normalized']), 'exceptions': plan['exceptions']}


def build_stakeholder_plan(workspace, content, filename):
    rows = _rows_from_file(content, filename, 'Stakeholders', STAKEHOLDER_COLUMNS)
    exceptions, normalized = [], []
    projects = {project.name.strip().lower(): project for project in Project.objects.filter(workspace=workspace)}
    seen = set()
    for row_number, values in enumerate(rows, start=2):
        project_name = str(values.get('project') or '').strip()
        name = str(values.get('name') or '').strip()
        project = projects.get(project_name.lower())
        if project is None:
            exceptions.append({'row': row_number, 'field': 'project', 'message': f'Project "{project_name}" was not found in this workspace.'})
            continue
        if not name or len(name) > 160:
            exceptions.append({'row': row_number, 'field': 'name', 'message': 'Stakeholder name is required and must be 160 characters or fewer.'})
            continue
        email = str(values.get('email') or '').strip()
        key = (project.id, (email.lower() if email else name.lower()))
        if key in seen:
            exceptions.append({'row': row_number, 'field': 'email' if email else 'name', 'message': 'Duplicate stakeholder in file.'})
            continue
        seen.add(key)
        influence = str(values.get('influence') or 'medium').strip().lower()
        interest = str(values.get('interest') or 'medium').strip().lower()
        if influence not in dict(ProjectStakeholder.INFLUENCE_CHOICES) or interest not in dict(ProjectStakeholder.INTEREST_CHOICES):
            exceptions.append({'row': row_number, 'field': 'influence', 'message': 'Influence and interest must be low, medium, or high.'})
            continue
        match_query = {'project': project}
        match_query['email__iexact' if email else 'name__iexact'] = email or name
        existing = ProjectStakeholder.objects.filter(**match_query).first()
        normalized.append({'row': row_number, 'action': 'update' if existing else 'create', 'existing_id': existing.id if existing else None, 'project_id': project.id, 'name': name, 'role': str(values.get('role') or '').strip(), 'email': email, 'influence': influence, 'interest': interest, 'notes': str(values.get('notes') or '').strip()})
    return {'kind': 'stakeholders', 'summary': {'total_rows': len(rows), 'creates': sum(item['action'] == 'create' for item in normalized), 'updates': sum(item['action'] == 'update' for item in normalized), 'exceptions': len(exceptions)}, 'normalized': normalized, 'exceptions': exceptions}


def commit_stakeholder_plan(workspace, actor, plan):
    with transaction.atomic():
        for item in plan['normalized']:
            stakeholder = ProjectStakeholder.objects.filter(id=item['existing_id'], project__workspace=workspace).first() if item['existing_id'] else ProjectStakeholder(project_id=item['project_id'])
            if stakeholder is None:
                raise ValueError('Stakeholder no longer exists in this workspace.')
            for field in ('project_id', 'name', 'role', 'email', 'influence', 'interest', 'notes'):
                setattr(stakeholder, field, item[field])
            stakeholder.full_clean()
            stakeholder.save()
    return {'created': sum(item['action'] == 'create' for item in plan['normalized']), 'updated': sum(item['action'] == 'update' for item in plan['normalized']), 'exceptions': plan['exceptions']}
