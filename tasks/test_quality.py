"""Quality-engineer tests: reporting calculations, drill-down filters,
permission scoping, reminder dedup, integrity checks and Excel import.

Kept separate from ``tasks/tests.py`` (which exercises the HTTP surface owned by
another agent) so these can change independently of view wiring.
"""

from datetime import date, datetime, timedelta
from io import BytesIO
import json

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from .automation import deliver_once, run_workspace_automation
from .integrity import run_integrity_checks
from .models import (
    AuditLog,
    ImportRun,
    LookupValue,
    Membership,
    NotificationDelivery,
    Project,
    ProjectStakeholder,
    Task,
    TaskChangeHistory,
    TaskCodeRegistry,
    TaskSubtask,
    Workspace,
    WorkspaceInvitation,
    WorkspaceNotification,
    WorkspaceSetting,
)
from .reporting import (
    apply_report_period,
    apply_task_filter,
    build_report,
    compute_kpis,
    project_health,
    scope_queryset,
    task_progress,
)


def make_workspace(slug='quality'):
    return Workspace.objects.create(name='Quality Workspace', slug=slug)


def make_user(username, email, first='', last=''):
    return User.objects.create_user(username=username, email=email, password='secure-pass-123', first_name=first, last_name=last)


def make_member(workspace, user, role='member'):
    return Membership.objects.create(workspace=workspace, user=user, role=role)


class ReportingCalculationTests(TestCase):
    def setUp(self):
        self.workspace = make_workspace()
        self.owner = make_user('owner@example.com', 'owner@example.com', 'Own', 'Er')
        make_member(self.workspace, self.owner, 'owner')

    def test_task_progress_precedence(self):
        done = Task.objects.create(workspace=self.workspace, title='Done', status='done', progress_percent=0)
        self.assertEqual(task_progress(done), 100)

        explicit = Task.objects.create(workspace=self.workspace, title='Explicit', status='in_progress', progress_percent=40)
        self.assertEqual(task_progress(explicit), 40)

        # progress_percent=0 is falsy -> falls back to the status mapping (in_progress=50).
        mapped = Task.objects.create(workspace=self.workspace, title='Mapped', status='in_progress', progress_percent=0)
        self.assertEqual(task_progress(mapped), 50)

        subtask_task = Task.objects.create(workspace=self.workspace, title='Subtasks', status='in_progress', progress_percent=0)
        self.assertEqual(task_progress(subtask_task, {subtask_task.id: (4, 1)}), 25)

    def test_build_report_totals(self):
        today = date(2026, 1, 15)
        project = Project.objects.create(workspace=self.workspace, name='Alpha')
        completed = Task.objects.create(workspace=self.workspace, title='C', project_ref=project, assignee=self.owner, status='done', progress_percent=100, completed_at=timezone.make_aware(datetime(2026, 1, 10, 12, 0)))
        active = Task.objects.create(workspace=self.workspace, title='A', project_ref=project, assignee=self.owner, status='in_progress', progress_percent=50, due_date=today + timedelta(days=3))
        overdue = Task.objects.create(workspace=self.workspace, title='O', status='todo', due_date=today - timedelta(days=1))
        blocked = Task.objects.create(workspace=self.workspace, title='B', assignee=self.owner, status='blocked', blocker_details='waiting')
        on_hold = Task.objects.create(workspace=self.workspace, title='H', assignee=self.owner, status='on_hold')
        cancelled = Task.objects.create(workspace=self.workspace, title='X', status='cancelled')
        unassigned = Task.objects.create(workspace=self.workspace, title='U', status='todo')

        report = build_report(self.workspace.id, scope='all', today=today)

        self.assertEqual(report['totals']['total_tasks'], 7)
        self.assertEqual(report['totals']['applicable_tasks'], 6)  # excludes cancelled
        self.assertEqual(report['totals']['cancelled_tasks'], 1)
        self.assertEqual(report['totals']['completed_tasks'], 1)
        self.assertEqual(report['totals']['completion_rate'], 17)  # 1 / 6
        self.assertEqual(report['overdue']['count'], 1)
        self.assertEqual(report['due_soon']['count'], 1)  # active, due in 3 days
        self.assertEqual(report['blocked']['count'], 1)
        self.assertEqual(report['on_hold']['count'], 1)
        self.assertEqual(report['unassigned']['count'], 2)  # overdue + unassigned (both todo, no assignee)

    def test_average_progress_uses_heuristics(self):
        today = date(2026, 1, 15)
        Task.objects.create(workspace=self.workspace, title='Done', status='done', progress_percent=100, completed_at=timezone.now())
        Task.objects.create(workspace=self.workspace, title='Half', status='in_progress', progress_percent=50)
        report = build_report(self.workspace.id, today=today)
        # (100 + 50) / 2 = 75
        self.assertEqual(report['totals']['average_progress'], 75)

    def test_scope_filters(self):
        project = Project.objects.create(workspace=self.workspace, name='Alpha')
        project_task = Task.objects.create(workspace=self.workspace, title='In project', project_ref=project)
        ops_task = Task.objects.create(workspace=self.workspace, title='Operations')

        self.assertEqual(set(scope_queryset(self.workspace.id, 'all').values_list('id', flat=True)), {project_task.id, ops_task.id})
        self.assertEqual(set(scope_queryset(self.workspace.id, 'operations').values_list('id', flat=True)), {ops_task.id})
        self.assertEqual(set(scope_queryset(self.workspace.id, 'project', project.id).values_list('id', flat=True)), {project_task.id})

    def test_apply_task_filter_vocabulary(self):
        design = LookupValue.objects.create(workspace=self.workspace, kind='workstream', name='Design', slug='design')
        assignee = make_user('a@example.com', 'a@example.com')
        Task.objects.create(workspace=self.workspace, title='Via ref', workstream_ref=design)
        Task.objects.create(workspace=self.workspace, title='Via char', workstream='Design')
        Task.objects.create(workspace=self.workspace, title='Other', workstream='Ops')
        Task.objects.create(workspace=self.workspace, title='Unassigned', assignee=assignee)

        qs = scope_queryset(self.workspace.id, 'all')
        # workstream matches both the ref and the legacy charfield
        self.assertEqual(apply_task_filter(qs, {'workstream': 'Design'}).count(), 2)
        self.assertEqual(apply_task_filter(qs, {'workstream': 'Ops'}).count(), 1)
        self.assertEqual(apply_task_filter(qs, {'assignee_id': None}).count(), 3)
        self.assertEqual(apply_task_filter(qs, {'assignee_id': assignee.id}).count(), 1)

    def test_report_period_is_delivery_based(self):
        today = date(2026, 1, 31)
        start = date(2026, 1, 1)
        end = date(2026, 1, 31)

        in_completed = Task.objects.create(workspace=self.workspace, title='In done', status='done', progress_percent=100, completed_at=timezone.make_aware(datetime(2026, 1, 15, 12, 0)))
        in_open = Task.objects.create(workspace=self.workspace, title='In open', status='in_progress', due_date=date(2026, 1, 10))
        out_completed = Task.objects.create(workspace=self.workspace, title='Out done', status='done', progress_percent=100, completed_at=timezone.make_aware(datetime(2026, 2, 1, 12, 0)))
        out_open = Task.objects.create(workspace=self.workspace, title='Out open', status='in_progress', due_date=date(2026, 2, 1))

        qs = scope_queryset(self.workspace.id, 'all')
        filtered = apply_report_period(qs, 'custom', today=today, start=start, end=end)
        ids = set(filtered.values_list('id', flat=True))
        self.assertEqual(ids, {in_completed.id, in_open.id})

    def test_compute_kpis_zero_target_protection(self):
        kpis = compute_kpis(applicable=6, completed=3, overdue=0, blocked=4, stale=1, targets={
            'completion_rate': 80, 'overdue': 0, 'blocked': 0, 'stale': 5,
        })
        self.assertEqual(kpis['completion_rate']['actual'], 50)
        self.assertFalse(kpis['completion_rate']['met'])  # 50% < 80% target
        self.assertEqual(kpis['overdue']['met'], True)
        self.assertEqual(kpis['overdue']['score'], 100.0)  # zero actual, zero target
        self.assertEqual(kpis['blocked']['met'], False)
        self.assertEqual(kpis['blocked']['score'], 0.0)  # zero target, nonzero actual
        self.assertEqual(kpis['stale']['met'], True)

    def test_project_health_states(self):
        project = Project.objects.create(workspace=self.workspace, name='Alpha', status='active', due_date=date(2026, 3, 1))
        today = date(2026, 1, 15)
        Task.objects.create(workspace=self.workspace, project_ref=project, title='On track', status='todo', due_date=date(2026, 2, 15))
        self.assertEqual(project_health(self.workspace.id, project, today=today)['health'], 'on-track')

        Task.objects.create(workspace=self.workspace, project_ref=project, title='Overdue', status='todo', due_date=today - timedelta(days=1))
        self.assertEqual(project_health(self.workspace.id, project, today=today)['health'], 'off-track')

    def test_progress_group_filters_are_replayable(self):
        project = Project.objects.create(workspace=self.workspace, name='Alpha')
        Task.objects.create(workspace=self.workspace, title='T', project_ref=project, status='todo')
        report = build_report(self.workspace.id, scope='all')
        project_group = next(g for g in report['progress_by_project'] if g['name'] == 'Alpha')
        self.assertEqual(project_group['filter'], {'project_id': project.id})

        # Replaying the emitted filter returns exactly the tasks in that group.
        qs = scope_queryset(self.workspace.id, 'all')
        self.assertEqual(apply_task_filter(qs, project_group['filter']).count(), project_group['total'])


class AutomationDedupTests(TestCase):
    def setUp(self):
        self.workspace = make_workspace(slug='dedup')
        self.leader = make_user('leader@example.com', 'leader@example.com')
        make_member(self.workspace, self.leader, 'owner')
        self.assignee = make_user('assignee@example.com', 'assignee@example.com')
        make_member(self.workspace, self.assignee, 'member')

    def test_deliver_once_is_idempotent(self):
        recipient = self.assignee
        first = deliver_once(self.workspace.id, recipient, 'due_soon_reminder', 'T', 'B', target_type='task', target_id=1, dedup_key='k:1')
        second = deliver_once(self.workspace.id, recipient, 'due_soon_reminder', 'T', 'B', target_type='task', target_id=1, dedup_key='k:1')
        self.assertTrue(first)
        self.assertFalse(second)
        self.assertEqual(NotificationDelivery.objects.filter(kind='due_soon_reminder').count(), 1)
        self.assertEqual(WorkspaceNotification.objects.filter(kind='due_soon_reminder').count(), 1)

    def test_run_workspace_automation_deduplicates(self):
        today = timezone.localdate()
        Task.objects.create(
            workspace=self.workspace, title='Due soon', assignee=self.assignee,
            status='in_progress', due_date=today + timedelta(days=1), state='active',
        )
        run_workspace_automation(self.workspace.id)
        run_workspace_automation(self.workspace.id)

        self.assertEqual(NotificationDelivery.objects.filter(kind='due_soon_reminder').count(), 1)
        self.assertEqual(WorkspaceNotification.objects.filter(kind='due_soon_reminder').count(), 1)

    def test_blocked_alert_notifies_owner_and_leaders_once(self):
        Task.objects.create(
            workspace=self.workspace, title='Blocked', assignee=self.assignee,
            status='blocked', blocker_details='dependency', state='active',
        )
        counts = run_workspace_automation(self.workspace.id)
        # owner (leader) + assignee, but assignee != leader -> two distinct recipients
        self.assertEqual(counts['blocked'], 2)
        self.assertEqual(NotificationDelivery.objects.filter(kind='blocked_alert').count(), 2)


class IntegrityCheckTests(TestCase):
    def setUp(self):
        self.workspace = make_workspace(slug='integrity')
        self.member = make_user('m@example.com', 'm@example.com')
        make_member(self.workspace, self.member, 'member')

    def _keys(self):
        return {result['key']: result for result in run_integrity_checks(self.workspace.id)}

    def test_duplicate_codes(self):
        # The DB constraint is case-sensitive, so a case variant can slip past
        # it while the integrity check (case-insensitive) still flags it.
        Task.objects.create(workspace=self.workspace, title='A', code='T-1')
        Task.objects.create(workspace=self.workspace, title='B', code='t-1')
        self.assertEqual(self._keys()['duplicate_task_codes']['count'], 2)

    def test_active_tasks_without_owners(self):
        Task.objects.create(workspace=self.workspace, title='No owner', status='todo', state='active')
        Task.objects.create(workspace=self.workspace, title='Done', status='done', progress_percent=100, completed_at=timezone.now())
        self.assertEqual(self._keys()['active_tasks_without_owners']['count'], 1)

    def test_multiple_active_owners(self):
        task = Task.objects.create(workspace=self.workspace, title='Owned', assignee=self.member, status='todo', state='active')
        task.supporters.add(self.member)  # same person -> discarded, not "multiple"
        self.assertEqual(self._keys()['multiple_active_owners']['count'], 0)

        other = make_user('o@example.com', 'o@example.com')
        make_member(self.workspace, other, 'member')
        task.supporters.add(other)
        self.assertEqual(self._keys()['multiple_active_owners']['count'], 1)

    def test_invalid_date_ranges(self):
        Task.objects.create(workspace=self.workspace, title='Bad', status='todo', start_date=date(2026, 2, 1), due_date=date(2026, 1, 1))
        self.assertEqual(self._keys()['invalid_date_ranges']['count'], 1)

    def test_completed_progress_inconsistencies(self):
        Task.objects.create(workspace=self.workspace, title='Done no completion', status='done', progress_percent=100)
        Task.objects.create(workspace=self.workspace, title='Done complete', status='done', progress_percent=100, completed_at=timezone.now())
        self.assertEqual(self._keys()['completed_progress_inconsistencies']['count'], 1)

    def test_orphan_project_relationships(self):
        other_workspace = make_workspace(slug='other')
        project = Project.objects.create(workspace=other_workspace, name='Foreign')
        Task.objects.create(workspace=self.workspace, title='Orphan', project_ref=project)
        self.assertEqual(self._keys()['orphan_project_relationships']['count'], 1)

    def test_incorrect_supporter_owner_relationships(self):
        non_member = make_user('outsider@example.com', 'outsider@example.com')
        task = Task.objects.create(workspace=self.workspace, title='Task', assignee=self.member)
        task.supporters.add(non_member)
        self.assertEqual(self._keys()['incorrect_supporter_owner_relationships']['count'], 1)


class ImportTests(TestCase):
    def setUp(self):
        self.workspace = make_workspace(slug='import')
        self.owner = make_user('owner@example.com', 'owner@example.com', 'Own', 'Er')
        make_member(self.workspace, self.owner, 'owner')
        self.member = make_user('alice@example.com', 'alice@example.com', 'Alice', 'Smith')
        make_member(self.workspace, self.member, 'member')

    def _workbook(self, rows, lists=None, owner_sheets=None):
        from openpyxl import Workbook

        wb = Workbook()
        general = wb.active
        general.title = 'General'
        general.append(list(self._header_map().values()))
        for row in rows:
            general.append(row)

        if lists:
            lists_ws = wb.create_sheet('Lists')
            for row in lists:
                lists_ws.append(row)

        for name, sheet_rows in (owner_sheets or {}).items():
            ws = wb.create_sheet(name)
            for row in sheet_rows:
                ws.append(row)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _header_map(self):
        from .importer import DEFAULT_GENERAL_COLUMNS
        return DEFAULT_GENERAL_COLUMNS

    def test_preview_is_read_only(self):
        buf = self._workbook([
            ['T-1', 'Imported task', '', 'alice@example.com', '', '', '', '', '', 'normal', 'todo', '', '2026-02-01', '', ''],
        ])
        from .importer import build_import_plan
        plan = build_import_plan(self.workspace, buf)
        self.assertEqual(plan['summary']['creates'], 1)
        self.assertEqual(plan['summary']['exceptions'], 0)
        self.assertEqual(Task.objects.count(), 0)  # nothing written
        self.assertEqual(ImportRun.objects.count(), 0)

    def test_commit_creates_and_updates_transactionally(self):
        existing = Task.objects.create(workspace=self.workspace, title='Will update', code='T-1')
        buf = self._workbook([
            ['T-1', 'Updated title', '', 'alice@example.com', '', '', '', '', '', 'high', 'in_progress', '', '2026-02-01', '', ''],
            ['T-2', 'New task', '', 'owner@example.com', '', '', '', '', '', 'normal', 'todo', '', '', '', ''],
        ])
        from .importer import build_import_plan, commit_import_plan
        plan = build_import_plan(self.workspace, buf)
        result = commit_import_plan(self.workspace, self.owner, plan)

        self.assertEqual(result['created'], 1)
        self.assertEqual(result['updated'], 1)
        existing.refresh_from_db()
        self.assertEqual(existing.title, 'Updated title')
        self.assertEqual(existing.priority, 'high')
        self.assertEqual(existing.status, 'in_progress')
        self.assertEqual(existing.assignee, self.member)
        self.assertEqual(Task.objects.filter(code='T-2').count(), 1)
        self.assertEqual(ImportRun.objects.filter(mode='commit').count(), 1)

    def test_unmatched_email_becomes_invitation(self):
        buf = self._workbook([
            ['T-9', 'Needs invite', '', 'newperson@example.com', '', '', '', '', '', 'normal', 'todo', '', '', '', ''],
        ])
        from .importer import build_import_plan, commit_import_plan
        plan = build_import_plan(self.workspace, buf)
        self.assertEqual(plan['summary']['invitations'], 1)
        commit_import_plan(self.workspace, self.owner, plan)
        self.assertTrue(WorkspaceInvitation.objects.filter(email='newperson@example.com', status='pending').exists())

    def test_unmatched_name_is_an_exception(self):
        buf = self._workbook([
            ['T-10', 'Ambiguous owner', '', 'Some Unknown Name', '', '', '', '', '', 'normal', 'todo', '', '', '', ''],
        ])
        from .importer import build_import_plan
        plan = build_import_plan(self.workspace, buf)
        self.assertEqual(plan['summary']['exceptions'], 1)
        self.assertEqual(plan['exceptions'][0]['field'], 'owner')

    def test_unknown_status_is_an_exception(self):
        buf = self._workbook([
            ['T-11', 'Bad status', '', 'alice@example.com', '', '', '', '', '', 'normal', 'wat', '', '', '', ''],
        ])
        from .importer import build_import_plan
        plan = build_import_plan(self.workspace, buf)
        self.assertEqual(plan['summary']['exceptions'], 1)
        self.assertEqual(plan['exceptions'][0]['field'], 'status')

    def test_owner_sheet_does_not_create_tasks(self):
        existing = Task.objects.create(workspace=self.workspace, title='Enrich me', code='T-20', status='in_progress', progress_percent=10)
        buf = self._workbook(
            [['T-20', 'Enrich me', '', 'alice@example.com', '', '', '', '', '', 'normal', 'in_progress', '', '', '', '']],
            owner_sheets={
                'Alice': [
                    ['Code', 'Progress %', 'Status', 'Blocker details', 'Actual completion'],
                    ['T-20', 80, 'in_progress', 'on hold', ''],
                    ['T-999', 50, 'done', '', ''],
                ],
            },
        )
        from .importer import build_import_plan, commit_import_plan
        plan = build_import_plan(self.workspace, buf)
        commit_import_plan(self.workspace, self.owner, plan)
        existing.refresh_from_db()
        self.assertEqual(existing.progress_percent, 80)
        self.assertEqual(existing.blocker_details, 'on hold')
        self.assertFalse(Task.objects.filter(code='T-999').exists())
        self.assertEqual(Task.objects.count(), 1)  # only the pre-existing task


class QualityHttpApiTests(TestCase):
    def setUp(self):
        self.workspace = make_workspace(slug='quality-http')
        self.owner = make_user('http-owner@example.com', 'http-owner@example.com', 'HTTP', 'Owner')
        self.member = make_user('http-member@example.com', 'http-member@example.com', 'HTTP', 'Member')
        self.outsider = make_user('http-outsider@example.com', 'http-outsider@example.com')
        make_member(self.workspace, self.owner, 'owner')
        make_member(self.workspace, self.member, 'member')
        self.project = Project.objects.create(workspace=self.workspace, name='API Project')

    def _workbook_bytes(self):
        from openpyxl import Workbook
        from .importer import DEFAULT_GENERAL_COLUMNS
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'General'
        sheet.append(list(DEFAULT_GENERAL_COLUMNS.values()))
        sheet.append(['IMP-1', 'Imported over HTTP', '', 'http-member@example.com', '', '', '', '', '', 'normal', 'todo', '', '', '', ''])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_reports_are_member_scoped_and_legacy_denominator_excludes_cancelled(self):
        Task.objects.create(workspace=self.workspace, title='Done', status='done', progress_percent=100, completed_at=timezone.now())
        Task.objects.create(workspace=self.workspace, title='Cancelled', status='cancelled')
        self.client.force_login(self.member)
        report = self.client.get(reverse('workspace-report', args=[self.workspace.id]))
        self.assertEqual(report.status_code, 200)
        self.assertEqual(report.json()['report']['totals']['completion_rate'], 100)
        legacy = self.client.get(reverse('report-summary', args=[self.workspace.id]))
        self.assertEqual(legacy.status_code, 200)
        self.assertEqual(legacy.json()['summary']['completion_rate'], 100)

        self.client.force_login(self.outsider)
        self.assertEqual(self.client.get(reverse('workspace-report', args=[self.workspace.id])).status_code, 403)

    def test_project_health_validates_workspace_project(self):
        self.client.force_login(self.member)
        response = self.client.get(reverse('project-health-report', args=[self.workspace.id]), {'project_id': self.project.id})
        self.assertEqual(response.status_code, 200)
        self.assertIn(response.json()['health']['health'], {'on-track', 'at-risk', 'off-track', 'completed'})
        other = make_workspace(slug='quality-other-http')
        foreign_project = Project.objects.create(workspace=other, name='Foreign')
        self.assertEqual(self.client.get(reverse('project-health-report', args=[self.workspace.id]), {'project_id': foreign_project.id}).status_code, 404)

    def test_integrity_and_automation_require_leader(self):
        self.client.force_login(self.member)
        self.assertEqual(self.client.get(reverse('workspace-integrity', args=[self.workspace.id])).status_code, 403)
        self.assertEqual(self.client.post(reverse('workspace-automation-run', args=[self.workspace.id])).status_code, 403)
        self.client.force_login(self.owner)
        self.assertEqual(self.client.get(reverse('workspace-integrity', args=[self.workspace.id])).status_code, 200)
        automation = self.client.post(reverse('workspace-automation-run', args=[self.workspace.id]))
        self.assertEqual(automation.status_code, 200)
        self.assertTrue(AuditLog.objects.filter(workspace=self.workspace, actor=self.owner, action='workspace_automation_run').exists())

    def test_import_requires_leader_and_exact_preview_checksum(self):
        content = self._workbook_bytes()
        preview_url = reverse('import-preview', args=[self.workspace.id])
        commit_url = reverse('import-commit', args=[self.workspace.id])
        self.client.force_login(self.member)
        denied = self.client.post(preview_url, {'workbook': SimpleUploadedFile('tasks.xlsx', content)})
        self.assertEqual(denied.status_code, 200)
        self.assertEqual(denied.json()['preview']['summary']['creates'], 1)

        self.client.force_login(self.owner)
        preview = self.client.post(preview_url, {'workbook': SimpleUploadedFile('tasks.xlsx', content)})
        self.assertEqual(preview.status_code, 200)
        checksum = preview.json()['preview']['checksum']
        preview_id = preview.json()['preview']['preview_id']
        self.assertEqual(preview.json()['preview']['summary']['creates'], 1)
        self.assertFalse(Task.objects.filter(code='IMP-1').exists())
        missing_checksum = self.client.post(commit_url, {'workbook': SimpleUploadedFile('tasks.xlsx', content)})
        self.assertEqual(missing_checksum.status_code, 409)
        committed = self.client.post(commit_url, {'workbook': SimpleUploadedFile('tasks.xlsx', content), 'preview_checksum': checksum, 'preview_id': preview_id})
        self.assertEqual(committed.status_code, 200)
        task = Task.objects.get(code='IMP-1')
        self.assertTrue(TaskCodeRegistry.objects.filter(workspace=self.workspace, code='IMP-1', task_id=task.id).exists())
        self.assertTrue(TaskChangeHistory.objects.filter(task=task, actor=self.owner, field='imported').exists())
        reused = self.client.post(commit_url, {'workbook': SimpleUploadedFile('tasks.xlsx', content), 'preview_checksum': checksum, 'preview_id': preview_id})
        self.assertEqual(reused.status_code, 409)

    def test_import_rejects_previously_used_task_code(self):
        TaskCodeRegistry.objects.create(workspace=self.workspace, code='IMP-1', task_id=99999)
        self.client.force_login(self.owner)
        preview = self.client.post(reverse('import-preview', args=[self.workspace.id]), {'workbook': SimpleUploadedFile('tasks.xlsx', self._workbook_bytes())})
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.json()['preview']['summary']['exceptions'], 1)
        self.assertIn('previously used', preview.json()['preview']['exceptions'][0]['message'])

    def test_project_and_stakeholder_csv_imports_and_templates(self):
        self.client.force_login(self.member)
        template = self.client.get(reverse('import-template', args=[self.workspace.id, 'projects', 'csv']))
        self.assertEqual(template.status_code, 200)
        self.assertIn('text/csv', template['Content-Type'])
        self.assertIn('workspace-projects-import-template.csv', template['Content-Disposition'])

        projects_csv = 'Name,Description,Status,Start date,End date,Due date,Timezone,Week anchor date,Due soon days,Configuration JSON\nApollo,Launch project,active,2026-01-01,2026-03-31,2026-03-31,Europe/London,2026-01-05,7,"{ ""stream"": ""A"" }"\n'.encode()
        preview_url = reverse('import-preview', args=[self.workspace.id])
        commit_url = reverse('import-commit', args=[self.workspace.id])
        project_preview = self.client.post(preview_url, {'import_type': 'projects', 'workbook': SimpleUploadedFile('projects.csv', projects_csv)})
        self.assertEqual(project_preview.status_code, 200)
        self.assertEqual(project_preview.json()['preview']['summary']['creates'], 1)
        self.client.force_login(self.owner)
        project_data = project_preview.json()['preview']
        project_commit = self.client.post(commit_url, {'import_type': 'projects', 'workbook': SimpleUploadedFile('projects.csv', projects_csv), 'preview_id': project_data['preview_id'], 'preview_checksum': project_data['checksum']})
        self.assertEqual(project_commit.status_code, 200, project_commit.content)
        project = Project.objects.get(name='Apollo')
        self.assertEqual(project.configuration, {'stream': 'A'})

        stakeholders_csv = b'Project,Name,Role,Email,Influence,Interest,Notes\nApollo,Pat Sponsor,Sponsor,pat@example.com,high,medium,Key approver\n'
        stakeholder_preview = self.client.post(preview_url, {'import_type': 'stakeholders', 'workbook': SimpleUploadedFile('stakeholders.csv', stakeholders_csv)})
        self.assertEqual(stakeholder_preview.status_code, 200)
        stakeholder_data = stakeholder_preview.json()['preview']
        stakeholder_commit = self.client.post(commit_url, {'import_type': 'stakeholders', 'workbook': SimpleUploadedFile('stakeholders.csv', stakeholders_csv), 'preview_id': stakeholder_data['preview_id'], 'preview_checksum': stakeholder_data['checksum']})
        self.assertEqual(stakeholder_commit.status_code, 200)
        self.assertTrue(ProjectStakeholder.objects.filter(project=project, email='pat@example.com', influence='high').exists())

    def test_csv_task_import_and_xlsx_template(self):
        self.client.force_login(self.owner)
        template = self.client.get(reverse('import-template', args=[self.workspace.id, 'tasks', 'xlsx']))
        self.assertEqual(template.status_code, 200)
        self.assertIn('spreadsheetml', template['Content-Type'])
        tasks_csv = b'Code,Title,Description,Owner,Supporters,Project,Workstream,Phase,Bucket,Priority,Status,Start date,Due date,Progress %,Labels,Blocker details,Actual completion\nCSV-1,CSV task,,,,,,,Backlog,normal,todo,,,0,,,\n'
        preview = self.client.post(reverse('import-preview', args=[self.workspace.id]), {'import_type': 'tasks', 'workbook': SimpleUploadedFile('tasks.csv', tasks_csv)})
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.json()['preview']['summary']['creates'], 1)
